from decimal import Decimal
from io import BytesIO

from django.contrib.staticfiles import finders
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

CINEGEST_NAVY = colors.HexColor("#071B3A")
CINEGEST_BLUE = colors.HexColor("#1565C0")
CINEGEST_YELLOW = colors.HexColor("#FFC400")
CINEGEST_TEXT = colors.HexColor("#172033")
CINEGEST_MUTED = colors.HexColor("#667085")
CINEGEST_BORDER = colors.HexColor("#D9E2F0")


def build_excel(rentas):
    """Genera un libro XLSX en memoria con las rentas suministradas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rentas"

    headers = ["ID", "Fecha", "Cliente", "Empleado", "Estado", "Total"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="071B3A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for renta in rentas:
        ws.append(
            [
                renta.id,
                timezone.localtime(renta.fecha_renta).strftime("%Y-%m-%d %H:%M"),
                renta.cliente.nombre,
                renta.empleado.nombre,
                renta.get_estado_display(),
                float(renta.total),
            ]
        )

    widths = {"A": 10, "B": 20, "C": 34, "D": 34, "E": 24, "F": 16}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _logo_path():
    return finders.find("images/cinegest-logo.png")


def _draw_page_header(canvas, doc):
    canvas.saveState()
    page_width, page_height = landscape(letter)

    canvas.setFillColor(CINEGEST_NAVY)
    canvas.rect(0, page_height - 54, page_width, 54, fill=1, stroke=0)
    canvas.setFillColor(CINEGEST_YELLOW)
    canvas.rect(0, page_height - 58, page_width, 4, fill=1, stroke=0)

    logo_path = _logo_path()
    if logo_path:
        logo = ImageReader(logo_path)
        image_width, image_height = logo.getSize()
        scale = min(150 / image_width, 42 / image_height)
        canvas.drawImage(
            logo,
            doc.leftMargin,
            page_height - 49,
            width=image_width * scale,
            height=image_height * scale,
            preserveAspectRatio=True,
            mask="auto",
        )
    else:
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(doc.leftMargin, page_height - 32, "CineGest")

    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(
        page_width - doc.rightMargin,
        page_height - 32,
        "Sistema de Gestión para Video Club",
    )

    canvas.setStrokeColor(CINEGEST_BORDER)
    canvas.line(doc.leftMargin, 30, page_width - doc.rightMargin, 30)
    canvas.setFillColor(CINEGEST_MUTED)
    canvas.setFont("Helvetica", 7.5)
    canvas.drawString(doc.leftMargin, 18, "CineGest · Reporte generado automáticamente")
    canvas.drawRightString(page_width - doc.rightMargin, 18, f"Página {doc.page}")
    canvas.restoreState()


def build_pdf(rentas, filtros):
    """Genera el reporte PDF corporativo de rentas y devuelve sus bytes."""
    rentas = list(rentas)
    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=72,
        bottomMargin=45,
        title="CineGest - Reporte de Rentas",
        author="CineGest",
        subject="Reporte operativo de rentas",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CineGestTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        textColor=CINEGEST_NAVY,
        spaceAfter=5,
    )
    subtitle_style = ParagraphStyle(
        "CineGestSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        leading=13,
        textColor=CINEGEST_MUTED,
        spaceAfter=14,
    )
    center_style = ParagraphStyle(
        "CineGestCenter",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=9,
        textColor=CINEGEST_MUTED,
    )
    right_style = ParagraphStyle(
        "CineGestRight",
        parent=styles["Normal"],
        alignment=TA_RIGHT,
        fontSize=8,
        textColor=CINEGEST_TEXT,
    )

    total_general = sum((renta.total for renta in rentas), Decimal("0.00"))
    generated_at = timezone.localtime().strftime("%d/%m/%Y %H:%M")

    story = [
        Paragraph("Reporte de Rentas", title_style),
        Paragraph(f"Resumen operativo generado el {generated_at}.", subtitle_style),
    ]

    summary = Table(
        [
            ["TOTAL DE RENTAS", "MONTO TOTAL", "DESDE", "HASTA", "ESTADO"],
            [
                str(len(rentas)),
                f"RD$ {total_general:,.2f}",
                filtros["desde"],
                filtros["hasta"],
                filtros["estado"],
            ],
        ],
        colWidths=[1.35 * inch, 1.65 * inch, 1.45 * inch, 1.45 * inch, 1.7 * inch],
    )
    summary.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), CINEGEST_NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BACKGROUND", (0, 1), (-1, 1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, 1), CINEGEST_TEXT),
                ("BOX", (0, 0), (-1, -1), 0.6, CINEGEST_BORDER),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, CINEGEST_BORDER),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend([summary, Spacer(1, 14)])

    if rentas:
        rows = [["#", "Fecha", "Cliente", "Empleado", "Estado", "Total"]]
        for renta in rentas:
            rows.append(
                [
                    f"#{renta.id}",
                    timezone.localtime(renta.fecha_renta).strftime("%d/%m/%Y %H:%M"),
                    renta.cliente.nombre,
                    renta.empleado.nombre,
                    renta.get_estado_display(),
                    Paragraph(f"RD$ {renta.total:,.2f}", right_style),
                ]
            )

        table = Table(
            rows,
            repeatRows=1,
            colWidths=[0.55 * inch, 1.25 * inch, 2.2 * inch, 2.2 * inch, 1.4 * inch, 1.1 * inch],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), CINEGEST_BLUE),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.35, CINEGEST_BORDER),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(table)
    else:
        story.append(Spacer(1, 24))
        story.append(Paragraph("No hay rentas que coincidan con los filtros aplicados.", center_style))

    doc.build(story, onFirstPage=_draw_page_header, onLaterPages=_draw_page_header)
    return stream.getvalue()
